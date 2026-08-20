from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from ._shared import *

# pandas/openpyxl 使用函数内惰性导入，避免环境中 numpy/cv2 版本冲突时
# 本模块连基本的姓氏提取和 @tool 定义都无法加载。


def _lazy_load_pandas():
    try:
        import pandas as pd  # type: ignore
        return pd
    except Exception:
        return None


def _lazy_load_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except Exception:
        return None


def _extract_surname(full_name: str) -> str:
    """从中文姓名中提取姓氏。

    支持常见复姓和单姓。
    """
    if not full_name:
        return ""
    full_name = full_name.strip()
    if len(full_name) == 0:
        return ""

    compound_surnames = [
        "欧阳", "太史", "端木", "上官", "司马", "东方", "独孤", "南宫",
        "万俟", "闻人", "夏侯", "诸葛", "尉迟", "公羊", "赫连", "澹台",
        "皇甫", "宗政", "濮阳", "公冶", "太叔", "申屠", "公孙", "慕容",
        "仲孙", "钟离", "长孙", "宇文", "司徒", "鲜于", "司空", "闾丘",
        "子车", "亓官", "司寇", "巫马", "公西", "颛孙", "壤驷", "公良",
        "漆雕", "乐正", "宰父", "谷梁", "拓跋", "夹谷", "轩辕", "令狐",
        "段干", "百里", "呼延", "东郭", "南门", "羊舌", "微生", "公户",
        "公玉", "公仪", "梁丘", "公仲", "公上", "公门", "公山", "公坚",
        "左丘", "公伯", "西门", "公祖", "第五", "公乘", "贯丘", "公皙",
        "南荣", "东里", "东宫", "仲长", "子书", "子桑", "即墨", "达奚",
        "褚师", "吴铭",
    ]
    if len(full_name) >= 2:
        first_two = full_name[:2]
        if first_two in compound_surnames:
            return first_two
    return full_name[0]


def _extract_given_name(full_name: str) -> str:
    """从中文姓名中提取名字（去掉姓氏）。"""
    if not full_name:
        return ""
    full_name = full_name.strip()
    surname = _extract_surname(full_name)
    return full_name[len(surname):] if len(full_name) > len(surname) else ""


@tool
def read_email_excel(excel_path: str, name_column: str = "姓名", email_column: str = "邮箱") -> str:
    """读取Excel文件中的收件人列表，提取姓名、邮箱等信息，并自动解析姓氏和名字。

    Excel应至少包含姓名和邮箱两列，其他列也会被保留并可在邮件模板中作为变量使用。

    Args:
        excel_path: Excel文件路径 (.xlsx / .xls)，例如 "/Users/xxx/contacts.xlsx"。
            文件需包含表头行。
        name_column: 姓名所在的列名，默认为"姓名"。
        email_column: 邮箱所在的列名，默认为"邮箱"。

    Returns:
        JSON字符串，包含 recipients 列表，每条记录包含：
        - name: 完整姓名
        - surname: 姓氏（自动提取，支持复姓）
        - given_name: 名字（去掉姓氏的部分）
        - email: 邮箱地址
        - 其他Excel中的原始列
    """
    try:
        resolved_input = Path(excel_path).resolve(strict=False)
        if not resolved_input.exists():
            return json.dumps({
                "status": "error",
                "message": f"Excel文件不存在: {excel_path}",
            }, ensure_ascii=False)

        suffix = resolved_input.suffix.lower()
        rows: list[dict[str, Any]] = []

        pd = _lazy_load_pandas()
        openpyxl = _lazy_load_openpyxl()

        if pd is not None:
            if suffix == ".xlsx":
                df = pd.read_excel(str(resolved_input), engine="openpyxl")
            elif suffix == ".xls":
                df = pd.read_excel(str(resolved_input), engine="xlrd")
            else:
                df = pd.read_excel(str(resolved_input))
            df = df.fillna("")
            for _, row in df.iterrows():
                row_dict = {str(k): ("" if pd.isna(v) else str(v).strip()) for k, v in row.to_dict().items()}
                rows.append(row_dict)
        elif openpyxl is not None and suffix in (".xlsx", ".xlsm"):
            wb = openpyxl.load_workbook(str(resolved_input), data_only=True)
            ws = wb.active
            headers: list[str] = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value is not None else "")
            for row_idx in range(2, ws.max_row + 1):
                row_dict: dict[str, Any] = {}
                for col_idx, header in enumerate(headers):
                    if not header:
                        continue
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    val = cell.value
                    row_dict[header] = "" if val is None else str(val).strip()
                if any(row_dict.values()):
                    rows.append(row_dict)
        else:
            return json.dumps({
                "status": "error",
                "message": "缺少Excel处理依赖，请先安装: pip install pandas openpyxl",
            }, ensure_ascii=False)

        if not rows:
            return json.dumps({
                "status": "error",
                "message": "Excel文件为空或没有数据行",
            }, ensure_ascii=False)

        available_columns = list(rows[0].keys())
        if name_column not in available_columns:
            return json.dumps({
                "status": "error",
                "message": f"找不到姓名列 '{name_column}'，可用列: {available_columns}",
            }, ensure_ascii=False)
        if email_column not in available_columns:
            return json.dumps({
                "status": "error",
                "message": f"找不到邮箱列 '{email_column}'，可用列: {available_columns}",
            }, ensure_ascii=False)

        recipients: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        for idx, row in enumerate(rows, start=1):
            raw_name = str(row.get(name_column, "")).strip()
            raw_email = str(row.get(email_column, "")).strip()
            if not raw_email:
                skipped.append({"row": idx, "reason": "邮箱为空", "data": row})
                continue
            if "@" not in raw_email:
                skipped.append({"row": idx, "reason": f"邮箱格式错误: {raw_email}", "data": row})
                continue
            record: dict[str, Any] = {
                **row,
                "name": raw_name,
                "surname": _extract_surname(raw_name),
                "given_name": _extract_given_name(raw_name),
                "email": raw_email,
            }
            recipients.append(record)

        logger.info(
            "📧 Excel读取: %s -> 成功%d条, 跳过%d条",
            str(resolved_input),
            len(recipients),
            len(skipped),
        )
        return json.dumps({
            "status": "success",
            "path": str(resolved_input),
            "total": len(recipients) + len(skipped),
            "recipients_count": len(recipients),
            "skipped_count": len(skipped),
            "available_columns": available_columns,
            "recipients": recipients,
            "skipped": skipped,
        }, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.exception("读取Excel失败")
        return json.dumps({
            "status": "error",
            "message": f"读取Excel出错: {e}",
        }, ensure_ascii=False)
