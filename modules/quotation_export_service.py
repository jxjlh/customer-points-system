import pandas as pd
import os
import json
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict, List


class QuotationExportService:
    def __init__(self):
        self.template_path = None
        self.template_mapping = {}
        self._load_config()

    def _load_config(self):
        config_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config")
        template_path = os.path.join(config_dir, "template_mapping.json")
        
        if os.path.exists(template_path):
            with open(template_path, 'r', encoding='utf-8') as f:
                self.template_mapping = json.load(f)

    def set_template_path(self, template_path: str) -> None:
        self.template_path = template_path

    def export_quotation(self, quotation_data: Dict, output_path: str = None) -> str:
        if self.template_path and os.path.exists(self.template_path):
            return self._export_with_template(quotation_data, output_path)
        else:
            return self._export_without_template(quotation_data, output_path)

    def _export_with_template(self, quotation_data: Dict, output_path: str = None) -> str:
        wb = load_workbook(self.template_path)
        ws = wb.active

        customer_info = quotation_data.get("customer_info", {})
        items = quotation_data.get("items", [])
        summary = quotation_data.get("summary", {})
        quote_number = quotation_data.get("quote_number", "")
        quote_date = quotation_data.get("quote_date", datetime.now().strftime("%Y-%m-%d"))

        customer_mapping = self.template_mapping.get("customer_info", {})
        for key, pos in customer_mapping.items():
            if key == "customer_name":
                self._write_cell(ws, pos["row"], pos["col"], customer_info.get("customer_name", ""))
            elif key == "contact_person":
                self._write_cell(ws, pos["row"], pos["col"], customer_info.get("contact_person", ""))
            elif key == "sales_person":
                self._write_cell(ws, pos["row"], pos["col"], customer_info.get("sales_person", ""))
            elif key == "customer_type":
                type_label = self._get_customer_type_label(customer_info.get("customer_type", ""))
                self._write_cell(ws, pos["row"], pos["col"], type_label)
            elif key == "quote_date":
                self._write_cell(ws, pos["row"], pos["col"], quote_date)
            elif key == "quote_number":
                self._write_cell(ws, pos["row"], pos["col"], quote_number)

        items_mapping = self.template_mapping.get("quote_items", {})
        start_row = items_mapping.get("start_row", 15)
        columns_map = items_mapping.get("columns", {})

        for i, item in enumerate(items):
            row = start_row + i
            
            for col_key, col_idx in columns_map.items():
                cell_value = self._get_item_value(item, col_key)
                col_letter = get_column_letter(col_idx + 1)
                cell = ws[f"{col_letter}{row}"]
                cell.value = cell_value

        summary_mapping = self.template_mapping.get("summary", {})
        for key, pos in summary_mapping.items():
            value = summary.get(key, 0)
            self._write_cell(ws, pos["row"], pos["col"], f"{value:,.2f}")

        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            return output_path
        else:
            export_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")
            os.makedirs(export_dir, exist_ok=True)
            file_name = f"{quote_number}.xlsx" if quote_number else f"报价单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(export_dir, file_name)
            wb.save(output_path)
            return output_path

    def _export_without_template(self, quotation_data: Dict, output_path: str = None) -> str:
        customer_info = quotation_data.get("customer_info", {})
        items = quotation_data.get("items", [])
        summary = quotation_data.get("summary", {})
        quote_number = quotation_data.get("quote_number", "")
        quote_date = quotation_data.get("quote_date", datetime.now().strftime("%Y-%m-%d"))

        wb = load_workbook(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "JAX小鼠活鼠报价单-CT-202605XX-GSJ00X-TB-甘模板.xlsx"))
        ws = wb.active

        ws.cell(row=6, column=2, value=customer_info.get("customer_name", ""))
        ws.cell(row=7, column=2, value=customer_info.get("contact_person", ""))
        ws.cell(row=8, column=2, value=customer_info.get("sales_person", ""))
        
        type_label = self._get_customer_type_label(customer_info.get("customer_type", ""))
        ws.cell(row=9, column=2, value=type_label)
        
        ws.cell(row=10, column=2, value=quote_date)
        ws.cell(row=11, column=2, value=quote_number)

        start_row = 15
        for i, item in enumerate(items):
            row = start_row + i
            ws.cell(row=row, column=1, value=item.get("strain", ""))
            ws.cell(row=row, column=2, value=item.get("name", ""))
            ws.cell(row=row, column=3, value=item.get("genotype", ""))
            ws.cell(row=row, column=4, value=item.get("age", ""))
            ws.cell(row=row, column=5, value=item.get("sex", ""))
            ws.cell(row=row, column=6, value=item.get("qty", 0))
            ws.cell(row=row, column=7, value=item.get("unit_price", 0))
            ws.cell(row=row, column=8, value=item.get("amount", 0))

        ws.cell(row=35, column=8, value=f"{summary.get('subtotal', 0):,.2f}")
        ws.cell(row=36, column=8, value=f"{summary.get('shipping', 0):,.2f}")
        ws.cell(row=37, column=8, value=f"{summary.get('service_fee', 0):,.2f}")
        ws.cell(row=38, column=8, value=f"{summary.get('discount', 0):,.2f}")
        ws.cell(row=39, column=8, value=f"{summary.get('tax', 0):,.2f}")
        ws.cell(row=40, column=8, value=f"{summary.get('grand_total', 0):,.2f}")

        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            return output_path
        else:
            export_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")
            os.makedirs(export_dir, exist_ok=True)
            file_name = f"{quote_number}.xlsx" if quote_number else f"报价单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(export_dir, file_name)
            wb.save(output_path)
            return output_path

    def _write_cell(self, ws, row: int, col: int, value: str) -> None:
        col_letter = get_column_letter(col)
        ws[f"{col_letter}{row}"].value = value

    def _get_item_value(self, item: Dict, key: str) -> str:
        if key == "strain":
            return item.get("strain", "")
        elif key == "name":
            return item.get("name", "")
        elif key == "genotype":
            return item.get("genotype", "")
        elif key == "age":
            return item.get("age", "")
        elif key == "sex":
            return item.get("sex", "")
        elif key == "qty":
            return item.get("qty", 0)
        elif key == "unit_price":
            return item.get("unit_price", 0)
        elif key == "amount":
            return item.get("amount", 0)
        return ""

    def _get_customer_type_label(self, customer_type: str) -> str:
        type_map = {
            "commercial": "Commercial",
            "npo": "NPO",
            "ka": "KA"
        }
        return type_map.get(customer_type, customer_type)

    def export_to_buffer(self, quotation_data: Dict) -> BytesIO:
        buffer = BytesIO()
        
        if self.template_path and os.path.exists(self.template_path):
            wb = load_workbook(self.template_path)
        else:
            template_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                                        "JAX小鼠活鼠报价单-CT-202605XX-GSJ00X-TB-甘模板.xlsx")
            if os.path.exists(template_file):
                wb = load_workbook(template_file)
            else:
                wb = load_workbook(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                                               "JAX小鼠活鼠报价单-CT-202605XX-GSJ00X-TB-甘模板.xlsx"))
        
        ws = wb.active

        customer_info = quotation_data.get("customer_info", {})
        items = quotation_data.get("items", [])
        summary = quotation_data.get("summary", {})
        quote_number = quotation_data.get("quote_number", "")
        quote_date = quotation_data.get("quote_date", datetime.now().strftime("%Y-%m-%d"))

        ws.cell(row=6, column=2, value=customer_info.get("customer_name", ""))
        ws.cell(row=7, column=2, value=customer_info.get("contact_person", ""))
        ws.cell(row=8, column=2, value=customer_info.get("sales_person", ""))
        
        type_label = self._get_customer_type_label(customer_info.get("customer_type", ""))
        ws.cell(row=9, column=2, value=type_label)
        
        ws.cell(row=10, column=2, value=quote_date)
        ws.cell(row=11, column=2, value=quote_number)

        start_row = 15
        for i, item in enumerate(items):
            row = start_row + i
            ws.cell(row=row, column=1, value=item.get("strain", ""))
            ws.cell(row=row, column=2, value=item.get("name", ""))
            ws.cell(row=row, column=3, value=item.get("genotype", ""))
            ws.cell(row=row, column=4, value=item.get("age", ""))
            ws.cell(row=row, column=5, value=item.get("sex", ""))
            ws.cell(row=row, column=6, value=item.get("qty", 0))
            ws.cell(row=row, column=7, value=item.get("unit_price", 0))
            ws.cell(row=row, column=8, value=item.get("amount", 0))

        ws.cell(row=35, column=8, value=f"{summary.get('subtotal', 0):,.2f}")
        ws.cell(row=36, column=8, value=f"{summary.get('shipping', 0):,.2f}")
        ws.cell(row=37, column=8, value=f"{summary.get('service_fee', 0):,.2f}")
        ws.cell(row=38, column=8, value=f"{summary.get('discount', 0):,.2f}")
        ws.cell(row=39, column=8, value=f"{summary.get('tax', 0):,.2f}")
        ws.cell(row=40, column=8, value=f"{summary.get('grand_total', 0):,.2f}")

        wb.save(buffer)
        buffer.seek(0)
        
        return buffer